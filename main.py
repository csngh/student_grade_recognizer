'''
Student Marksheet Scanner v3.0 (Static Dataset Prototype)
Created: 23/09/2016
Updated: 21/03/2017
Author: Chetan Singh

'''

import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from difflib import SequenceMatcher
from PIL import Image
import pytesseract
import openpyxl
import xlrd
import csv

student_dict = {1: ['RAVI', 'PATEL', 'TS111111'], 2: ['MANI', 'SINGH', 'TS111112'], 3: ['GAGAN', 'DEEP', 'TS111113'], 
                4: ['VICKY', 'SINGH', 'TS111114'], 5: ['AJAY', 'PARMAR', 'TS111115']}

#Compare Scanned Data with Database
def compare_data(fName, lName):
    for i in student_dict.keys():
        if(SequenceMatcher(None, fName.upper(), student_dict[i][0]).ratio() >= 0.7):
            gotFname = student_dict[i][0]
            for j in student_dict.keys():
                if(SequenceMatcher(None, lName.upper(), student_dict[j][1]).ratio() >= 0.7):
                    gotLname = student_dict[j][1]
                    break
                else:
                    gotLname = 'No match!'
            break
        else:
            gotFname = 'No match!'
            gotID = ''
            
    for k in student_dict.keys():
        if (student_dict[k][0] == gotFname.upper() and student_dict[k][1] == gotLname.upper()):
            gotID = student_dict[k][2]
            break
        else:
            gotID = 'No match!' 
    
    return [gotFname, gotLname, gotID]

#Write all Headers in a New Sheet
def write_all_headers(sheet):
    print('Writing all headers...')
    sheet.cell(row = 1, column = 1).value = 'Student ID'
    sheet.cell(row = 1, column = 2).value = 'Student Name'
    colId = 3
    
    for i in range(1, 6):
        for q in 'a b c d e'.split():
            sheet.cell(row = 1, column = colId).value = 'Q' + str(i) + '(' + q + ')'
            colId += 1        
    sheet.cell(1, 28).value = 'Total'

#Write all Headers in a Sheet for Portal
def write_headers_excel_portal(sheet):
    print('Writing all headers for portal file...')
    sheet.cell(row = 1, column = 1).value = 'OrgDefinedDID'
    sheet.cell(row = 1, column = 2).value = 'First Name'
    sheet.cell(row = 1, column = 3).value = 'Last Name'
    sheet.cell(row = 1, column = 4).value = 'Test 1 Points Grade'
    sheet.cell(row = 1, column = 5).value = 'End-of-Line-Indicator'
    
#Write All Details
def write_all_details(sheet, rowID, dataList):
    print("Writing student's data and generating sheet to upload on portal...\n")
    colID = 1
    for data in dataList:
        sheet.cell(row = rowID, column = colID).value = data
        colID += 1

#Convert XLSX to CSV
def convert_to_csv(fileName):
    workbook_csv = xlrd.open_workbook(fileName + '.xlsx')
    sheet_csv = workbook_csv.sheet_by_index(0)
    csv_file = open(fileName + '.csv', 'w')
    csv_writer = csv.writer(csv_file, quoting=csv.QUOTE_ALL)

    for rownum in range(sheet_csv.nrows):
        csv_writer.writerow(sheet_csv.row_values(rownum))
    csv_file.close()

def run_main(fileName):
    #Converting image to text format
    text = pytesseract.image_to_string(Image.open(fileName), lang = 'eng')

    #Saving all in a list
    extracted_text = text.split()

    #Fetching Name
    nameStartIndex = extracted_text.index('Name:') + 1
    nameEndIndex = nameStartIndex + 2
    studentName = ''

    for i in range(nameStartIndex, nameEndIndex):
        studentName += extracted_text[i]
        studentName += ' '
        
    #Fetching Student ID
    #idStartIndex = extracted_text.index('ID:') + 1
    #studentID = extracted_text[idStartIndex]

    #Comparing Details
    student_details = compare_data(studentName.split()[0], studentName.split()[1])
    print(student_details)

    #Marks in Question 1
    q1StartIndex = extracted_text.index('Marks') + 2
    q1Marks = []

    for i in range(q1StartIndex, q1StartIndex + 5):
        q1Marks.append(extracted_text[i])

    #Marks in Question 2
    q2StartIndex = q1StartIndex + 7
    q2Marks = []

    for i in range(q2StartIndex, q2StartIndex + 5):
        q2Marks.append(extracted_text[i])
        
    #Marks in Question 3
    q3StartIndex = q2StartIndex + 7
    q3Marks = []

    for i in range(q3StartIndex, q3StartIndex + 5):
        q3Marks.append(extracted_text[i])
        
    #Marks in Question 4
    q4StartIndex = q3StartIndex + 7
    q4Marks = []

    for i in range(q4StartIndex, q4StartIndex + 5):
        q4Marks.append(extracted_text[i])
        
    #Marks in Question 5
    q5StartIndex = q4StartIndex + 7
    q5Marks = []

    for i in range(q5StartIndex, q5StartIndex + 5):
        q5Marks.append(extracted_text[i])
        
    #Total Marks
    totalStartIndex = extracted_text.index('Total') + 1
    totalMarks = extracted_text[totalStartIndex]

    #Preparing the Excel Sheet
    try:
        workbook = openpyxl.load_workbook('marksList.xlsx')
        sheet = workbook.active
        
        #If Sheet is Blank
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(row = 1, column = 1).value == None:
            write_all_headers(sheet)
            #Writing Data to Sheet
            details_to_write = [student_details[2]] + [student_details[0] + ' ' + student_details[1]] + q1Marks + q2Marks + q3Marks + q4Marks + q5Marks + [totalMarks]
            write_all_details(sheet, sheet.max_row + 2, details_to_write)
        #If Sheet has some Values
        else:
            rowID = sheet.max_row
            #colID = sheet.max_column
            #Writing Data to Sheet
            details_to_write = [student_details[2]] + [student_details[0] + ' ' + student_details[1]] + q1Marks + q2Marks + q3Marks + q4Marks + q5Marks + [totalMarks]
            write_all_details(sheet, rowID + 1, details_to_write)
    except:
        print("Excel file markslist doesn't exists, creating new!")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        write_all_headers(sheet)
        #Writing Data to Sheet
        details_to_write = [student_details[2]] + [student_details[0] + ' ' + student_details[1]] + q1Marks + q2Marks + q3Marks + q4Marks + q5Marks + [totalMarks]
        write_all_details(sheet, 2, details_to_write)

    #Preparing the Portal Acceptable Excel Sheet
    try:
        workbookPortal = openpyxl.load_workbook('marksToUpload.xlsx')
        sheetPortal = workbookPortal.active
        
        #If Sheet is Blank
        if sheetPortal.max_row == 1 and sheetPortal.max_column == 1 and sheetPortal.cell(row = 1, column = 1).value == None:
            write_headers_excel_portal(sheetPortal)
            #Writing Data to Sheet
            details_to_write = [student_details[2]] + [student_details[0], student_details[1]] + [totalMarks] + ['#']
            write_all_details(sheetPortal, sheetPortal.max_row + 2, details_to_write)
        #If Sheet has some Values
        else:
            rowID = sheetPortal.max_row
            #colID = sheetPortal.max_column
            #Writing Data to Sheet
            details_to_write = [student_details[2]] + [student_details[0], student_details[1]] + [totalMarks] + ['#']
            write_all_details(sheetPortal, rowID + 1, details_to_write)
    except:
        print("Excel file marksToUpload doesn't exists, creating new!")
        workbookPortal = openpyxl.Workbook()
        sheetPortal = workbookPortal.active
        write_headers_excel_portal(sheetPortal)
        #Writing Data to Sheet
        details_to_write = [student_details[2]] + [student_details[0], student_details[1]] + [totalMarks] + ['#']
        write_all_details(sheetPortal, 2, details_to_write)

    #Saving the Sheet
    workbook.save('marksList.xlsx')
    workbookPortal.save('marksToUpload.xlsx')
    convert_to_csv('marksList')
    convert_to_csv('marksToUpload')

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(MainWindow.sizePolicy().hasHeightForWidth())
        MainWindow.setSizePolicy(sizePolicy)
        MainWindow.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("image/browse.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        MainWindow.setWindowIcon(icon)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout.setObjectName("gridLayout")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setObjectName("label")
        self.verticalLayout.addWidget(self.label)
        self.gridLayout.addLayout(self.verticalLayout, 0, 0, 1, 1)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.imagePath = QtWidgets.QLineEdit(self.centralwidget)
        self.imagePath.setObjectName("imagePath")
        self.horizontalLayout.addWidget(self.imagePath)
        self.browseButton = QtWidgets.QPushButton(self.centralwidget)
        self.browseButton.setObjectName("browseButton")
        self.horizontalLayout.addWidget(self.browseButton)
        self.gridLayout.addLayout(self.horizontalLayout, 1, 0, 1, 1)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.progressBar = QtWidgets.QProgressBar(self.centralwidget)
        self.progressBar.setProperty("value", 0)
        self.progressBar.setObjectName("progressBar")
        self.horizontalLayout_2.addWidget(self.progressBar)
        self.convertButton = QtWidgets.QPushButton(self.centralwidget)
        self.convertButton.setObjectName("convertButton")
        self.horizontalLayout_2.addWidget(self.convertButton)
        self.gridLayout.addLayout(self.horizontalLayout_2, 2, 0, 1, 1)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 496, 20))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Marks Scanner"))
        self.label.setText(_translate("MainWindow", "Select Image File to Scan (.jpg recommended)"))
        self.imagePath.setText(_translate("MainWindow", "Image Path Appears Here."))
        self.browseButton.setText(_translate("MainWindow", "Browse"))
        self.convertButton.setText(_translate("MainWindow", "Scan and Convert"))

        self.convertButton.clicked.connect(self.run)
        self.browseButton.clicked.connect(self.openFileNameDialog)

    def run(self):
        self.progressCompleted = 0
        while self.progressCompleted < 100:
            self.progressCompleted += 0.0001
            self.progressBar.setValue(self.progressCompleted)
        if '.jpg' in self.imagePath.text() or '.jpeg' in self.imagePath.text():
            run_main(self.imagePath.text())
        else:
            self.imagePath.setText('Invalid Path!')

    def openFileNameDialog(self):    
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        fileName, _ = QtWidgets.QFileDialog.getOpenFileName(None, "Select Image to Scan", "","All Files (*);;JPEG Files (*.jpeg);;JPG Files (*.jpg)", options=options)
        if fileName:
            self.imagePath.setText(fileName)
        else:
            self.imagePath.setText('Invalid Path!')

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())